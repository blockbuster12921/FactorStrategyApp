import pandas as pd
import numpy as np
import logging
import openpyxl, openpyxl.utils.dataframe
from . import factor_calc, score_calc, factor_cluster
from .utils import MeasureTime, MeasureBlockTime

logger = logging.getLogger(__name__)

class FactorFilter():

    @staticmethod
    def create_state(cluster_df, min_cluster_count):

        clusters = sorted(cluster_df['Cluster'].unique())

        state = {
            'Complete': False,
            'BestStage': 0,
            'Stages': [
                {
                    'BaselineScore': None,
                    'Trials': [{ 'Cluster': int(cluster), 'ScoreDelta': None } for cluster in clusters],
                }
            ]
        }

        # Check number of clusters against min
        if len(clusters) <= min_cluster_count:
            state['Complete'] = True

        return state


    def __init__(
        self, 
        returns_data, 
        factors_data,
        cluster_df,
        st_duration,
        st_weight,
        objective,
        score_pairs_range,
        removal_fraction,
        min_cluster_count,
        state,
        ):

        self.returns_data = returns_data
        self.factors_data = factors_data
        self.cluster_df = cluster_df

        self.cluster_expected_returns = {}
        clusters = factor_cluster.get_clustered_factors_from_df(cluster_df)
        for i, cluster_factors in clusters.items():
            if len(cluster_factors) == 1:
                self.cluster_expected_returns[i] = self.factors_data[cluster_factors[0]]
            else:
                self.cluster_expected_returns[i] = factor_calc.combine_factors(self.factors_data, cluster_factors)

        assert(st_duration >= 0)
        assert(st_weight >= 0.0)
        self.st_weight = st_weight

        self.st_months = None
        self.st_period_start = None
        if (self.st_weight > 0.0) and (st_duration > 0):
            self.st_months = self.returns_data.columns.tolist()[-st_duration:]
            self.st_period_start = self.st_months[0]

        self.objective = objective

        self.score_pairs_range = score_pairs_range
 
        self.removal_fraction = removal_fraction
        self.min_cluster_count = min_cluster_count
 
        self.state = state

    def run(self):

        if self.state['Complete']:
            return

        # Calculate next score
        self._calc_next_score()

        # Update state
        self._update_state()

    def _update_state(self):

        # Check for a completed stage
        stage = self.state['Stages'][-1]

        if stage['BaselineScore'] is None:
            return

        for trial in stage['Trials']:
            if trial['ScoreDelta'] is None:
                return

        # Stage is complete

        # Update best stage
        if stage['BaselineScore'] > self.state['Stages'][self.state['BestStage']]['BaselineScore']:
            self.state['BestStage'] = len(self.state['Stages'])-1

        # Check against min
        remove_count = max(1, int(self.removal_fraction*len(stage['Trials'])))
        retain_count = len(stage['Trials']) - remove_count
        if retain_count <= self.min_cluster_count:
            self.state['Complete'] = True
            return

        # Add new stage with retained clusters
        score_deltas = pd.DataFrame(stage['Trials']).sort_values('ScoreDelta')
        retained = sorted(score_deltas.head(retain_count)['Cluster'].tolist())
        self.state['Stages'].append(
            {
                'BaselineScore': None,
                'Trials': [{'Cluster': int(cluster), 'ScoreDelta': None} for cluster in retained],
            }
        )

    def calc_combination_score(self, clusters):

        factor_av = factor_calc.combine_factors(self.cluster_expected_returns, clusters)

        if self.objective == 'score':

            deltas = factor_calc.calc_long_short_pair_return_deltas(factor_av, self.returns_data, self.score_pairs_range[0], self.score_pairs_range[1])
            metrics = factor_calc.calc_long_short_pairs_metrics(deltas)
            score = score_calc.calc_score(metrics)

            if self.st_weight > 0.0:
                st_metrics = factor_calc.calc_long_short_pairs_metrics(deltas, start_date=self.st_period_start)
                st_score = score_calc.calc_score(st_metrics)
                score = (score + self.st_weight*st_score)/(1.0 + self.st_weight)

        elif (self.objective == 'corr') or (self.objective == 'rank_corr'):

            corr_method = 'spearman' if self.objective == 'rank_corr' else 'pearson'

            corrs = factor_av.corrwith(self.returns_data, axis=0, method=corr_method)
            score = corrs.mean()

            if self.st_months is not None:
                st_score = corrs.loc[self.st_months].mean()
                score = (score + self.st_weight*st_score)/(1.0 + self.st_weight)

            score *= 100.0

        else:
            raise ValueError("Invalid objective method '{}'".format(self.objective))

        return score

    def _calc_next_score(self):

        # Construct list of clusters to evaluate
        stage = self.state['Stages'][-1]
        clusters = [trial['Cluster'] for trial in stage['Trials']]
        active_trial = None
        if stage['BaselineScore'] is not None:
            for trial in stage['Trials']:
                if trial['ScoreDelta'] is None:
                    active_trial = trial
                    clusters.remove(trial['Cluster'])
                    break
            assert(active_trial is not None)

        # Calculate score for cluster combination
        score = self.calc_combination_score(clusters)

        # Store score
        if stage['BaselineScore'] is None:
            stage['BaselineScore'] = score
        else:
            active_trial['ScoreDelta'] = score - stage['BaselineScore']


class FactorFilterReportGenerator():
    
    def __init__(self, db, project_id):
        
        self.db = db
        self.project_id = project_id

        self.bold_font = openpyxl.styles.Font(b=True)
        
    def generate(self):

        self.data_info = self.db.get_project_data_info(self.project_id)

        factor_info = { factor_index: factor_name for factor_index, factor_name in enumerate(self.data_info['Factors']) }
        factor_info[-1] = 'Baseline'
        self.factor_info_df = pd.Series(factor_info).to_frame('Factor').sort_index()

        factor_filter_list = self.db.get_project_factor_filter_state(self.project_id)
        factor_filter_list = sorted(factor_filter_list, key=lambda x: x['TargetMonth'])

        self.target_months = []
        self.factor_filter_df_list = []
        self.retained = {}

        for ff in factor_filter_list:
            self._process_target_month(ff)

        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)

        ws = wb.create_sheet('Retained')
        self._write_dataframe(ws, pd.DataFrame(self.retained).rename_axis('Factor').reset_index())

        for index, ff_df in enumerate(self.factor_filter_df_list):
            ws = wb.create_sheet("{:%b-%Y}".format(self.target_months[index]))
            self._write_dataframe(ws, ff_df)
            ws.column_dimensions['A'].width = 20

        return wb
    
    def _write_dataframe(self, ws, df):
        
        for r in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for col in range(1, len(df.columns)+1):
            ws.cell(row=1, column=col).font = self.bold_font
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = 15

        ws.auto_filter.ref = ws.dimensions

    def _process_target_month(self, ff):

        target_month = ff['TargetMonth']
        self.target_months.append(target_month)
        
        target_month_str = "{:%b-%Y}".format(target_month)

        cluster_df = self.db.get_project_factor_clusters(self.project_id, target_month)
        cluster_factors = factor_cluster.get_clustered_factors_from_df(cluster_df)
        
        df = self.factor_info_df.copy()

        df.loc[-1][df.columns[0]] = 'Baseline'

        for stage_index, stage in enumerate(ff['Stages']):
            score_delta = pd.DataFrame(stage['Trials']).set_index('Cluster').astype(float)

            col = '{} Clusters'.format(len(stage['Trials']))
            df[col] = np.nan

            for trial in stage['Trials']:
                cluster = trial['Cluster']
                score_delta = trial['ScoreDelta']
                factors = cluster_factors[cluster]
                for f in factors:
                    df.loc[df.index.isin(factors), col] = score_delta

            df.at[-1, col] = stage['BaselineScore']

        self.factor_filter_df_list.append(df)

        retained_list = df[['Factor',df.columns[-1]]].loc[df.index != -1].set_index('Factor')
        retained_list = retained_list[retained_list.columns[0]]
        retained_list = retained_list.loc[retained_list.notnull()].index.tolist()
        self.retained[target_month_str] = { f: True for f in retained_list }

            


    








