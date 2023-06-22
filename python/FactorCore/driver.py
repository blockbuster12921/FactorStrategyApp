import pandas as pd
import numpy as np
import openpyxl, openpyxl.utils.dataframe
import scipy.stats
import logging
from . import factor_calc, score_calc, kernel_regression

logger = logging.getLogger(__name__)

class Driver():
    
    def __init__(
        self,
        method,
        returns_data,
        factors_data,
        forward_month_stocks_disabled,
        in_sample_date_range,
        output_date_range,
        factor_outlier_rejection_multiplier,
        ranked_returns_stddev_multiplier,
        smoothing_kernel="rbf",
        smoothing_bandwidth=100.0,
        correlation_average_method='median',
        correlation_lookback_months=120,
        correlation_basis='rolling_out_of_sample',
        calculation_settings=None,
    ):
        assert(method in ['fixed','rolling','unbiased','correlation','variable_direction','correlation_average_over_stocks','rank_average_correlation','rank_ttest','binned_rank_correlation'])
        self.method = method
        
        self.returns_data = returns_data.copy().dropna(how='all', axis=0)

        assert(len(factors_data.keys()) > 0)
        self.factors_data = factors_data

        # Append forward month to returns with zero placeholder
        forward_month = (self.returns_data.index[-1] + pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
        forward_month_returns = np.zeros(len(self.returns_data.columns))
        forward_month_returns[forward_month_stocks_disabled] = np.nan
        self.returns_data.loc[forward_month] = forward_month_returns

        self.in_sample_start_date = self.returns_data.index[0]
        if in_sample_date_range[0] is not None:
            self.in_sample_start_date = max(pd.Timestamp(in_sample_date_range[0]), self.in_sample_start_date)
        self.in_sample_end_date = pd.Timestamp(in_sample_date_range[1])
        
        self.output_start_date = pd.Timestamp(output_date_range[0])
        self.output_end_date = forward_month
        if output_date_range[1] is not None:
            self.output_end_date = min(pd.Timestamp(output_date_range[1]), self.output_end_date)

        assert(self.in_sample_start_date <= self.in_sample_end_date)
        assert(self.output_start_date <= self.output_end_date)

        logger.debug("In-sample date range: {} to {}".format(self.in_sample_start_date, self.in_sample_end_date))
        logger.debug("Output date range: {} to {}".format(self.output_start_date, self.output_end_date))

        self.output_months = [month for month in self.returns_data.index if \
                         (month >= self.output_start_date) and (month <= self.output_end_date)]

        self.returns_data = self.returns_data.loc[(self.returns_data.index >= self.in_sample_start_date) & \
                                                  (self.returns_data.index <= self.output_end_date)]
        self.returns_df_T = self.returns_data.T

        self.factor_outlier_rejection_multiplier = factor_outlier_rejection_multiplier
        self.ranked_returns_stddev_multiplier = ranked_returns_stddev_multiplier

        self.factors_expected_returns = {}
        self.factor_weights = {}

        self.smoothing_kernel = smoothing_kernel
        self.smoothing_bandwidth = smoothing_bandwidth

        self.correlation_average_method = correlation_average_method
        self.correlation_lookback_months = correlation_lookback_months
        self.correlation_basis = correlation_basis
        self.calculation_settings = calculation_settings

    def run(self):
        
        for factor_index in self.factors_data.keys():
            self._calc_factor_expected_returns(factor_index)

        self.factor_weights = pd.Series(self.factor_weights)

    def _calc_factor_expected_returns(self, factor_index):
                
        factor_df = self.factors_data[factor_index]

        factor_df = factor_df.loc[(factor_df.index >= self.in_sample_start_date) & \
                                  (factor_df.index <= self.output_end_date)]

        # Remove factor outliers
        factor_df = self._remove_factor_outliers(factor_df)
        
        # Set factor value to nan if there is no corresponding return
        factor_df += 0 * self.returns_data

        if self.method == 'variable_direction':
            self.factors_expected_returns[factor_index] = factor_df.rank(pct=True, axis=1).T[self.output_months]
            return

        if self.method == 'correlation':
            if (self.correlation_basis == 'rolling_out_of_sample') or (self.correlation_basis == 'rolling_in_sample'):
                self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_correlation_expected_ranks_rolling(factor_df)
            elif self.correlation_basis == 'all_in_sample':
                self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_correlation_expected_ranks_simple(factor_df)
            else:
                raise ValueError("Invalid correlation basis '{}'".format(self.correlation_basis))
            return

        if self.method == 'correlation_average_over_stocks':
            if (self.correlation_basis == 'rolling_out_of_sample') or (self.correlation_basis == 'rolling_in_sample'):
                self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_correlation_average_over_stocks_expected_ranks_rolling(factor_df)
            elif self.correlation_basis == 'all_in_sample':
                self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_correlation_average_over_stocks_expected_ranks_simple(factor_df)
            else:
                raise ValueError("Invalid correlation basis '{}'".format(self.correlation_basis))
            return

        if self.method == 'rank_average_correlation':
            self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_rank_average_correlation(factor_df)
            return

        if self.method == 'rank_ttest':
            self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_rank_ttest(factor_df)
            return

        if self.method == 'binned_rank_correlation':
            self.factors_expected_returns[factor_index], self.factor_weights[factor_index] = self._calc_binned_rank_correlation(factor_df)
            return


        if self.method == 'unbiased':
            # Convert factors to percentiles 
            f = factor_df.T.rank(pct=True)

            # Construct df with vectors of factors and returns for modelling
            X = self.returns_df_T.unstack().to_frame('r')
            X['f'] = f.unstack()

            idx = pd.IndexSlice
            X_train = X.loc[idx[self.in_sample_start_date:self.in_sample_end_date, :]]
            X_train = X_train.dropna()
            
            X_predict = X['f'].loc[idx[self.output_start_date:self.output_end_date, :]]
            X_predict = X_predict.dropna()

            smoother = kernel_regression.KernelRegression(kernel=self.smoothing_kernel, gamma=self.smoothing_bandwidth)
            smoother.fit(X_train['f'].values.reshape(-1, 1), X_train['r'].values)
            y = smoother.predict(X_predict.values.reshape(-1, 1))

            X_predict = X_predict.to_frame('f')
            X_predict['r'] = y

            # Construct expected returns df
            expected_returns = X_predict['r'].unstack(0)            
            expected_returns = expected_returns.reindex(factor_df.columns)

            self.factors_expected_returns[factor_index] = expected_returns
            return
        

        # Sort returns by factor value for each month
        factor_df = factor_df.T
        sorted_returns_df = self._calc_sorted_returns(factor_df)

        # Perform outlier rejection on sorted returns
        sorted_returns_df = self._remove_sorted_returns_outliers(sorted_returns_df)

        # Calculate factor ranks
        ranks = factor_df[self.output_months].rank(method='first', ascending=False) - 1.0

        if self.method == 'fixed':

            # Calculate average and count of ranked returns
            sorted_returns_average = sorted_returns_df.mean(axis=1)
            sorted_returns_average_values = sorted_returns_average.values

            # Assign average returns to stocks by factor value rank
            def rank_to_av_return(x):
                return np.nan if np.isnan(x) else sorted_returns_average_values[int(x)]

            expected_returns = ranks.applymap(rank_to_av_return)

        else:

            assert(self.method == 'rolling')
            
            # Make sure we have all required months before averaging
            for col in reversed(ranks.columns):
                if col in sorted_returns_df.columns:
                    break
                sorted_returns_df[col] = np.nan

            # Calculate expanding average and shift back by 1 month
            sorted_returns_average = sorted_returns_df.T.expanding(0).mean().shift(1)
            sorted_returns_average = sorted_returns_average.loc[self.output_start_date:].T
            
            # Calculate expected returns
            expected_returns = ranks.copy()
            for col in expected_returns.columns:
                col_av_returns = sorted_returns_average[col].values

                def rank_to_av_return(x):
                    return np.nan if np.isnan(x) else col_av_returns[int(x)]

                expected_returns[col] = ranks[col].map(rank_to_av_return).values

        self.factors_expected_returns[factor_index] = expected_returns

    def _calc_sorted_returns(self, factor_df):

        end_date = self.in_sample_end_date if self.method == 'fixed' else self.returns_df_T.columns[-1]

        fitting_months = [col for col in self.returns_df_T.columns if (col >= self.in_sample_start_date) and (col <= end_date)]
        sorted_returns_df = pd.DataFrame(index=range(0,len(factor_df.index)), columns=fitting_months, data=np.nan)

        for month in sorted_returns_df.columns:

            f = factor_df[month].values
            r = self.returns_df_T[month].values

            sorted_indexes = np.argsort(f, kind='stable') 
            sorted_returns = r[sorted_indexes]

            nan_count = np.sum(np.isnan(f))
            if nan_count > 0:
                sorted_returns = sorted_returns[:-nan_count]
                sorted_returns_df[month].values[:-nan_count] = sorted_returns[::-1]
            else:
                sorted_returns_df[month] = sorted_returns[::-1]

        return sorted_returns_df
        
    def _remove_factor_outliers(self, factor_df):
        
        if self.method in ['fixed','unbiased']:

            # Calculate mean and std over in-sample period
            in_sample_df = factor_df.loc[(factor_df.index >= self.in_sample_start_date) & \
                                       (factor_df.index <= self.in_sample_end_date)]

            in_sample_mean = np.nanmean(in_sample_df.values)
            in_sample_stddev = np.nanstd(in_sample_df.values, ddof=1)

            # Apply mask over entire factor df
            cond = factor_df.sub(in_sample_mean).abs().sub(self.factor_outlier_rejection_multiplier * in_sample_stddev).fillna(1.0)
            return factor_df.mask(cond > 0)
        
        else:

            # Calculate expanding mean and std for each month
            expanding = factor_df.stack(dropna=False).droplevel(1).expanding(0)
            col_count = len(factor_df.columns)
            stats = expanding.mean()[(col_count-1)::col_count].to_frame('mean')
            stats['std'] = expanding.std(ddof=1)[(col_count-1)::col_count]

            # Backfill using last month of in-sample period
            stats.loc[stats.index < self.in_sample_end_date] = np.nan
            stats = stats.bfill()

            # Calculate and apply mask
            cond = factor_df.sub(stats['mean'], axis=0).abs().sub(self.factor_outlier_rejection_multiplier * stats['std'], axis=0).fillna(1.0)
            return factor_df.mask(cond > 0)

    def _remove_sorted_returns_outliers(self, sorted_returns_df):
        
        if self.method == 'fixed':

            in_sample_months = [month for month in sorted_returns_df.columns if \
                 (month >= self.in_sample_start_date) and (month <= self.in_sample_end_date)]
            in_sample_df = sorted_returns_df[in_sample_months]

            in_sample_mean = in_sample_df.mean(axis=1)
            in_sample_stddev = in_sample_df.std(ddof=1, axis=1)

            cond = sorted_returns_df.sub(in_sample_mean, axis=0).abs().sub(self.ranked_returns_stddev_multiplier * in_sample_stddev, axis=0).fillna(1.0)
            return sorted_returns_df.mask(cond > 0)

        else:

            assert(self.method == 'rolling')

            sorted_returns_df = sorted_returns_df.T
            
            # Calculate expanding mean and std for each month
            expanding = sorted_returns_df.expanding(0)
            mean = expanding.mean()
            std = expanding.std(ddof=1)

            # Backfill using last month of in-sample period
            mean.loc[mean.index < self.in_sample_end_date] = np.nan
            mean = mean.bfill()
            std.loc[std.index < self.in_sample_end_date] = np.nan
            std = std.bfill()

            # Calculate and apply mask
            cond = sorted_returns_df.sub(mean).abs().sub(self.ranked_returns_stddev_multiplier * std).fillna(1.0)
            sorted_returns_df = sorted_returns_df.mask(cond > 0)

            return sorted_returns_df.T

    def _calc_correlation_expected_ranks_simple(self, factor_df):

        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        corr_end_date = (self.output_end_date - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
        corr_start_date = (corr_end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

        # Calculate rank correlations
        corrs = 100.0 * self.returns_data.loc[corr_start_date:corr_end_date].corrwith(factor_df.loc[corr_start_date:corr_end_date], axis=1, method='spearman')

        if self.correlation_average_method == 'median':
            weight = corrs.median()
        elif self.correlation_average_method == 'mean':
            weight = corrs.mean()
        else:
            raise ValueError("Invalid Correlation average method '{}'".format(self.correlation_average_method))

        if corrs.count() < min(self.calculation_settings['DriverCorrelationMinMonths'], self.correlation_lookback_months, len(corrs)):
            weight = np.nan

        # Apply sign of weight to factor values
        expected_ranks = np.sign(weight) * factor_df.loc[self.output_months]

        # Rank factor values
        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight > 0:
            # Scale ranks by magnitude of weight
            expected_ranks *= np.abs(weight)**scale_by_weight

        return expected_ranks.T, weight

    def _calc_correlation_expected_ranks_rolling(self, factor_df):

        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        # Calculate rank correlations
        corrs = 100.0 * self.returns_data.corrwith(factor_df, axis=1, method='spearman')

        # Calculate weights
        min_periods = min(self.calculation_settings['DriverCorrelationMinMonths'], self.correlation_lookback_months, len(corrs)-1)
        rolling = corrs.rolling(window=self.correlation_lookback_months, min_periods=min(min_periods, self.correlation_lookback_months))
        if self.correlation_average_method == 'median':
            weights = rolling.median()
        elif self.correlation_average_method == 'mean':
            weights = rolling.mean()
        else:
            raise ValueError("Invalid Correlation average method '{}'".format(self.correlation_average_method))

        if self.correlation_basis == 'rolling_out_of_sample':
            weights = weights.shift(1)
        else:
            # Use last in-sample month for forward (OOS) month
            weights.iloc[-1] = weights.iloc[-2]

        # Apply sign of weights to factor values
        expected_ranks = factor_df.mul(np.sign(weights), axis=0)

        # Rank factor values
        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight > 0:
            # Scale ranks by magnitude of weight
            expected_ranks = expected_ranks.mul(weights.abs()**scale_by_weight, axis=0)

        expected_ranks = expected_ranks.T[self.output_months]

        return expected_ranks, weights.iloc[-1]


    def _calc_correlation_average_over_stocks_expected_ranks_simple(self, factor_df):

        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        min_data_count = 12

        corr_end_date = (self.output_end_date - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
        corr_start_date = (corr_end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

        returns_ranks = self.returns_data.loc[corr_start_date:corr_end_date].rank(axis=1, pct=True)
        factor_ranks = factor_df.loc[returns_ranks.index].rank(axis=1, pct=True)

        corrs = returns_ranks.corrwith(factor_ranks, axis=0, method='pearson')
        corrs = corrs.mask(factor_ranks.count(axis=0) < min_data_count)

        if self.correlation_average_method == 'mean':
            weight = corrs.mean()
        elif self.correlation_average_method == 'median':
            weight = corrs.median()
        else:
            raise ValueError("Invalid correlation average method '{}'".format(self.correlation_average_method))

        # Apply sign of weight to factor values
        expected_ranks = np.sign(weight) * factor_df.loc[self.output_months]

        # Rank factor values
        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight > 0:
            # Scale ranks by magnitude of weight
            expected_ranks *= np.abs(weight)**scale_by_weight

        return expected_ranks.T, weight


    def _calc_correlation_average_over_stocks_expected_ranks_rolling(self, factor_df):

        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        def calc_rolling_factor_stock_correlation(returns_df, factor_df, start_date, end_date, min_data_count=12):

            data_start_date = (start_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)
            
            returns_df = self.returns_data.loc[data_start_date:end_date].rank(axis=1, pct=True)

            factor_df = factor_df.loc[returns_df.index].rank(axis=1, pct=True)

            corrs = {}
            
            target_months = [month for month in returns_df.index if (month >= start_date) and (month <= end_date)]
            
            for month in target_months:
                period_start = (month - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)
                period_factor_df = factor_df.loc[period_start:month]
                corrs[month] = returns_df.loc[period_start:month].corrwith(period_factor_df, axis=0, method='pearson')
                corrs[month] = corrs[month].mask(period_factor_df.count(axis=0) < min_data_count)

            corrs = pd.DataFrame(corrs)

            return corrs

        def calc_factor_stock_correlation_weights(returns_df, factor_df, start_date, end_date, min_data_count=12):

            shifted_start_date = (start_date - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
            
            corr_df = calc_rolling_factor_stock_correlation(
                returns_df, factor_df, shifted_start_date, end_date, min_data_count=min_data_count)
            
            if self.correlation_average_method == 'mean':
                weights = corr_df.mean()
            elif self.correlation_average_method == 'median':
                weights = corr_df.median()
            else:
                raise ValueError("Invalid correlation average method '{}'".format(self.correlation_average_method))

            if self.correlation_basis == 'rolling_out_of_sample':
                weights = weights.shift(1)
            else:
                # Use last in-sample month for forward (OOS) month
                weights.iloc[-1] = weights.iloc[-2]

            weights = weights.loc[start_date:]

            return weights

        weights = calc_factor_stock_correlation_weights(self.returns_data, factor_df, self.output_months[0], self.output_months[-1], min_data_count=12)

        expected_ranks = factor_df.loc[self.output_months].mul(np.sign(weights), axis=0)

        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight:
            # Scale ranks by magnitude of weight
            expected_ranks = expected_ranks.mul(weights.abs()**scale_by_weight, axis=0)

        return expected_ranks.T, weights.iloc[-1]


    def _calc_rank_average_correlation_factor_weight(self, factor_df, returns_df, start_date, end_date, min_data_months, average, sort_method, splice_frac):

        factor_df = factor_df.loc[start_date:end_date].T
        returns_df = returns_df.loc[start_date:end_date].T

        returns_df += 0*factor_df
        returns_df = returns_df.dropna(how='all', axis=1)

        factor_df += 0*returns_df
        factor_df = factor_df.dropna(how='all', axis=1)

        returns_pct_rank_df = (returns_df.rank(pct=False, axis=0, ascending=True)-1.0)
        returns_pct_rank_df /= returns_pct_rank_df.max()

        if sort_method == 'asc':
            sorted_returns_pct_rank_df = factor_calc.sort_returns_by_factor(
                factor_df, returns_pct_rank_df, ascending=True).dropna(how='all')
        elif sort_method == 'desc':
            sorted_returns_pct_rank_df = factor_calc.sort_returns_by_factor(
                factor_df, returns_pct_rank_df, ascending=False).dropna(how='all')
        else:
            assert(sort_method == 'spliced')

            sorted_returns_pct_rank_df = factor_calc.sort_returns_by_factor(
                factor_df, returns_pct_rank_df, ascending=True).dropna(how='all')

            sorted_returns_pct_rank_desc_df = factor_calc.sort_returns_by_factor(
                factor_df, returns_pct_rank_df, ascending=False).dropna(how='all')

            sorted_returns_pct_rank_desc_df.index = sorted_returns_pct_rank_desc_df.index[::-1]

            sorted_returns_pct_rank_df.loc[sorted_returns_pct_rank_df.index > len(sorted_returns_pct_rank_df)/2] = sorted_returns_pct_rank_desc_df

            if splice_frac < 1.0:
                assert(splice_frac > 0.0)
                splice_min = int(0.5*splice_frac*len(sorted_returns_pct_rank_df.index))
                splice_max = len(sorted_returns_pct_rank_df.index) - splice_min - 1
                sorted_returns_pct_rank_df = sorted_returns_pct_rank_df.loc[
                    (sorted_returns_pct_rank_df.index < splice_min) | 
                    (sorted_returns_pct_rank_df.index > splice_max)
                ].reset_index(drop=True)

        if min_data_months > 0:
            data_months = sorted_returns_pct_rank_df.count(axis=1)
            sorted_returns_pct_rank_df = sorted_returns_pct_rank_df.loc[data_months >= min_data_months]

        if average == 'median':
            average_sorted_returns_pct_rank = sorted_returns_pct_rank_df.median(axis=1)
        else:
            assert(average == 'mean')
            average_sorted_returns_pct_rank = sorted_returns_pct_rank_df.mean(axis=1)
        average_sorted_returns_pct_rank = average_sorted_returns_pct_rank.dropna()

        corr = np.corrcoef(average_sorted_returns_pct_rank, average_sorted_returns_pct_rank.index)[0][1]

        if sort_method == 'desc':
            corr = -corr

        return corr


    def _calc_rank_average_correlation(self, factor_df):

        min_data_months = int(self.calculation_settings['DriverReturnRankCorrelationMinDataMonths'])
        sort_method = self.calculation_settings['DriverReturnRankCorrelationOrder']
        splice_frac = float(self.calculation_settings['DriverReturnRankCorrelationSplicePercent'])/100.0
        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        weights = pd.Series(index=self.output_months, data=np.nan)

        if self.correlation_basis == 'all_in_sample':
            end_date = (self.output_end_date - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
            start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

            weights[:] = self._calc_rank_average_correlation_factor_weight(
                factor_df, self.returns_data.copy(), start_date, end_date, average=self.correlation_average_method, 
                min_data_months=min_data_months, sort_method=sort_method, splice_frac=splice_frac)

        elif self.correlation_basis == 'rolling_in_sample':

            for i, output_month in enumerate(self.output_months[:-1]):
                end_date = output_month
                start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

                weights.iloc[i] = self._calc_rank_average_correlation_factor_weight(
                    factor_df, self.returns_data.copy(), start_date, end_date, average=self.correlation_average_method, 
                    min_data_months=min_data_months, sort_method=sort_method, splice_frac=splice_frac)

            weights.iloc[-1] = weights.iloc[-2]

        elif self.correlation_basis == 'rolling_out_of_sample':

            for i, output_month in enumerate(self.output_months):
                end_date = (output_month - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
                start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

                weights.iloc[i] = self._calc_rank_average_correlation_factor_weight(
                    factor_df, self.returns_data.copy(), start_date, end_date, average=self.correlation_average_method, 
                    min_data_months=min_data_months, sort_method=sort_method, splice_frac=splice_frac)

        else:
            raise ValueError("Invalid Correlation Basis '{}'".format(self.correlation_basis))

        # Apply sign of weight to factor values
        expected_ranks = factor_df.loc[self.output_months].mul(np.sign(weights), axis=0)

        # Rank factor values
        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight > 0:
            # Scale ranks by magnitude of weight
            expected_ranks = expected_ranks.mul(weights.abs()**scale_by_weight, axis=0)

        return expected_ranks.T, weights.iloc[-1]


    def _calc_rank_ttest_factor_weight(self, factor_df, returns_df, start_date, end_date, prob_frac):

        factor_df = factor_df.loc[start_date:end_date].T
        returns_df = returns_df.loc[start_date:end_date].T

        returns_df += 0*factor_df

        returns_pct_rank_df = (returns_df.rank(pct=False, axis=0, ascending=True)-1.0)
        returns_pct_rank_df /= returns_pct_rank_df.max()

        factor_pct_rank_df = (factor_df.rank(pct=False, axis=0, ascending=True)-1.0)
        factor_pct_rank_df /= factor_pct_rank_df.max()

        df = pd.DataFrame(data={ 
                'f': factor_pct_rank_df.values.flatten(), 
                'r': returns_pct_rank_df.values.flatten() 
            })

        # Make sure some data is included at both ends
        p = max(prob_frac, df.f.min(), 1.0 - df.f.max())

        t_value, p_value = scipy.stats.ttest_ind(df.loc[df.f >= (1.0-p)].r, df.loc[df.f <= p].r)
        
        return t_value  


    def _calc_rank_ttest(self, factor_df):

        prob_frac = float(self.calculation_settings['DriverReturnRankTTestPercentage'])/100.0
        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        weights = pd.Series(index=self.output_months, data=np.nan)

        if self.correlation_basis == 'all_in_sample':
            end_date = (self.output_end_date - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
            start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

            weights[:] = self._calc_rank_ttest_factor_weight(
                factor_df, self.returns_data.copy(), start_date, end_date, prob_frac=prob_frac)

        elif self.correlation_basis == 'rolling_in_sample':

            for i, output_month in enumerate(self.output_months[:-1]):
                end_date = output_month
                start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

                weights.iloc[i] = self._calc_rank_ttest_factor_weight(
                    factor_df, self.returns_data.copy(), start_date, end_date, prob_frac=prob_frac)

            weights.iloc[-1] = weights.iloc[-2]

        elif self.correlation_basis == 'rolling_out_of_sample':

            for i, output_month in enumerate(self.output_months):
                end_date = (output_month - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
                start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

                weights.iloc[i] = self._calc_rank_ttest_factor_weight(
                    factor_df, self.returns_data.copy(), start_date, end_date, prob_frac=prob_frac)

        else:
            raise ValueError("Invalid Correlation Basis '{}'".format(self.correlation_basis))

        # Apply sign of weight to factor values
        expected_ranks = factor_df.loc[self.output_months].mul(np.sign(weights), axis=0)

        # Rank factor values
        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight > 0:
            # Scale ranks by magnitude of weight
            expected_ranks = expected_ranks.mul(weights.abs()**scale_by_weight, axis=0)

        return expected_ranks.T, weights.iloc[-1]


    def _calc_binned_rank_correlation_factor_weight(self, factor_df, returns_df, start_date, end_date, bin_frac):

        factor_df = factor_df.loc[start_date:end_date].T
        returns_df = returns_df.loc[start_date:end_date].T

        returns_df += 0*factor_df

        returns_pct_rank_df = (returns_df.rank(pct=False, axis=0, ascending=True)-1.0)
        returns_pct_rank_df /= returns_pct_rank_df.max()

        factor_pct_rank_df = (factor_df.rank(pct=False, axis=0, ascending=True)-1.0)
        factor_pct_rank_df /= factor_pct_rank_df.max()

        bins = int(len(factor_pct_rank_df)*bin_frac)

        df = pd.DataFrame(data={ 
                'f': factor_pct_rank_df.values.flatten(), 
                'r': returns_pct_rank_df.values.flatten() 
            })

        df['bin'] = (df.f * (bins-1)).round()

        binned_return_ranks = df.groupby('bin').r
        
        if self.correlation_average_method == 'median':
            binned_return_ranks = binned_return_ranks.median()
        else:
            assert(self.correlation_average_method == 'mean')
            binned_return_ranks = binned_return_ranks.mean()

        corr = np.corrcoef(binned_return_ranks.index, binned_return_ranks.values)[0][1]

        return corr

    def _calc_binned_rank_correlation(self, factor_df):

        bin_frac = float(self.calculation_settings['DriverBinnedReturnRankCorrelationBinPercentage'])/100.0
        scale_by_weight = float(self.calculation_settings['DriverFactorWeightScale'])

        weights = pd.Series(index=self.output_months, data=np.nan)

        if self.correlation_basis == 'all_in_sample':
            end_date = (self.output_end_date - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
            start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

            weights[:] = self._calc_binned_rank_correlation_factor_weight(
                factor_df, self.returns_data.copy(), start_date, end_date, bin_frac=bin_frac)

        elif self.correlation_basis == 'rolling_in_sample':

            for i, output_month in enumerate(self.output_months[:-1]):
                end_date = output_month
                start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

                weights.iloc[i] = self._calc_binned_rank_correlation_factor_weight(
                    factor_df, self.returns_data.copy(), start_date, end_date, bin_frac=bin_frac)

            weights.iloc[-1] = weights.iloc[-2]

        elif self.correlation_basis == 'rolling_out_of_sample':

            for i, output_month in enumerate(self.output_months):
                end_date = (output_month - pd.DateOffset(months=1)) + pd.offsets.MonthEnd(0)
                start_date = (end_date - pd.DateOffset(months=self.correlation_lookback_months-1)) + pd.offsets.MonthEnd(0)

                weights.iloc[i] = self._calc_binned_rank_correlation_factor_weight(
                    factor_df, self.returns_data.copy(), start_date, end_date, bin_frac=bin_frac)

        else:
            raise ValueError("Invalid Correlation Basis '{}'".format(self.correlation_basis))

        # Apply sign of weight to factor values
        expected_ranks = factor_df.loc[self.output_months].mul(np.sign(weights), axis=0)

        # Rank factor values
        expected_ranks = expected_ranks.rank(pct=True, axis=1)

        if scale_by_weight > 0:
            # Scale ranks by magnitude of weight
            expected_ranks = expected_ranks.mul(weights.abs()**scale_by_weight, axis=0)

        return expected_ranks.T, weights.iloc[-1]


class DriverParamSelector():

    def __init__(
        self, 
        returns_data, 
        factors_data, 
        driver_method,
        forward_month_stocks_disabled,
        in_sample_date_range,
        param_selection_date_range,
        factor_outlier_rejection_multiplier,
        ranked_returns_stddev_multiplier,
        correlation_average_method,
        correlation_lookback_months,
        correlation_basis,
        calculation_settings,
        score_pairs_range,
        param,
        state,
        param_values,
        ):
        self.returns_data = returns_data
        self.factors_data = factors_data
        self.driver_method = driver_method
        self.forward_month_stocks_disabled = forward_month_stocks_disabled
        self.in_sample_date_range = in_sample_date_range
        self.param_selection_date_range = param_selection_date_range
        self.factor_outlier_rejection_multiplier = factor_outlier_rejection_multiplier
        self.ranked_returns_stddev_multiplier = ranked_returns_stddev_multiplier
        self.correlation_average_method = correlation_average_method
        self.correlation_lookback_months = correlation_lookback_months
        self.correlation_basis = correlation_basis
        self.score_pairs_range = score_pairs_range
        self.calculation_settings = calculation_settings
        self.param = param
        self.state = state
        self.param_values = param_values

        # Initialise state if necessary
        if self.state is None:

            assert(param_values is not None)
            assert(len(param_values) > 0)

            self.state = {
                    'Complete': False,
                    'Trials': [{'ParamValue': param_value, 'Score': None} for param_value in param_values],
                }

    def run(self):

        if self.state['Complete']:
            return

        # Calculate next score
        self._calc_next_score()

        # Update state
        self._update_state()

    def _update_state(self):

        # Check for completion
        self.state['Complete'] = True
        for trial in self.state['Trials']:
            if trial['Score'] is None:
                self.state['Complete'] = False
                break

    def _calc_next_score(self):

        # Get next parameter value to try
        driver_factor_outlier_rejection_multiplier = self.factor_outlier_rejection_multiplier
        driver_ranked_returns_stddev_multiplier = self.ranked_returns_stddev_multiplier

        trial = None
        for trial in self.state['Trials']:
            if trial['Score'] is None:
                break
        assert(trial is not None)

        if self.param == 'FactorOutlierRejection':
            driver_factor_outlier_rejection_multiplier = trial['ParamValue']
        elif self.param == 'ReturnsOutlierRejection':
            driver_ranked_returns_stddev_multiplier = trial['ParamValue']
        else:
            assert(False)

        # Run driver
        driver_instance = Driver(
            method=self.driver_method, 
            returns_data=self.returns_data,
            factors_data=self.factors_data,
            forward_month_stocks_disabled=self.forward_month_stocks_disabled,
            in_sample_date_range=self.in_sample_date_range,
            output_date_range=self.param_selection_date_range,
            factor_outlier_rejection_multiplier=driver_factor_outlier_rejection_multiplier,
            ranked_returns_stddev_multiplier=driver_ranked_returns_stddev_multiplier,
            correlation_average_method=self.correlation_average_method,
            correlation_lookback_months=self.correlation_lookback_months,
            correlation_basis=self.correlation_basis,
            calculation_settings=self.calculation_settings,
            )

        driver_instance.run()

        # Combine factor expected returns
        factor_av = factor_calc.combine_factors(
            driver_instance.factors_expected_returns, 
            list(driver_instance.factors_expected_returns.keys())
            )

        # Calculate score
        returns_df = self.returns_data.loc[factor_av.columns].T
        deltas = factor_calc.calc_long_short_pair_return_deltas(
            factor_av, 
            returns_df, 
            self.score_pairs_range[0], self.score_pairs_range[1],
            )
        metrics = factor_calc.calc_long_short_pairs_metrics(deltas)
        trial['Score'] = score_calc.calc_score(metrics)


class DriverParamSelectionReportGenerator():

    def __init__(self, db, project_id):        
        self.db = db
        self.project_id = project_id

        self.bold_font = openpyxl.styles.Font(b=True)

    def generate(self, params):

        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)

        for param in params:
            ws = wb.create_sheet(param)
            self._generate_param(param, ws)

        return wb

    def _generate_param(self, param, ws):

        states = self.db.get_project_driver_param_selection_state_list(self.project_id, param)
        states = sorted(states, key=lambda x: x['TargetMonth'])
        
        dfs = []
        for state in states:
            df = pd.DataFrame(state['Trials'])
            df.columns = ['Parameter', "{:%b-%Y}".format(state['TargetMonth'])]
            df = df.set_index('Parameter')
            dfs.append(df)

        if len(dfs) == 0:
            return

        df = pd.concat(dfs, axis=1)
        df = df.reset_index()

        self._write_dataframe(ws, df)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _write_dataframe(self, ws, df):
        
        for r in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for col in range(1, len(df.columns)+1):
            ws.cell(row=1, column=col).font = self.bold_font
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = 15

        ws.auto_filter.ref = ws.dimensions


class FactorWeightsReportGenerator():

    def __init__(self, db, project_id):
        self.db = db
        self.project_id = project_id
    
    def _write_dataframe(self, ws, df):
        
        for r in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        ws.auto_filter.ref = ws.dimensions

    def generate(self):

        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)

        data_info = self.db.get_project_data_info(self.project_id)
        factor_names = data_info['Factors']

        df = self.db.get_project_driver_factor_weights(self.project_id)

        df.columns = ["{:%b-%y}".format(col) for col in df.columns]
        df.index.name = "Factor"
        df = df.reset_index()
        df['Factor'] = [factor_names[f] for f in df['Factor']]

        ws = wb.create_sheet('Data')
        self._write_dataframe(ws, df)

        return wb
