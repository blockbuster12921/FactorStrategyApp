import pandas as pd
import numpy as np
import logging
import openpyxl, openpyxl.utils.dataframe
from . import factor_calc, score_calc, driver
from .utils import MeasureTime, MeasureBlockTime

logger = logging.getLogger(__name__)

def get_indexes_for_disabled_stocks(stocks, disabled):
    if disabled is None:
        return []

    disabled_indexes = []
    for item in disabled:
        for i, stock in enumerate(stocks):
            if (stock['DatasetIndex'] == item['DatasetIndex']) and (stock['DatasetID'] == item['DatasetID']):
                disabled_indexes.append(i)
                break

    return disabled_indexes

class StockSelector():

    @staticmethod
    def create_state(stocks_enabled, min_stock_count, stocks_removed_by_market_cap_filter):

        assert(stocks_enabled is not None)
        assert(min_stock_count > 0)

        state = {
            'Complete': False,
            'MarketCapFiltered': stocks_removed_by_market_cap_filter,
            'Stages': [
                {
                    'StockRemoved': None,
                    'BaselineScore': None,
                    'Trials': [{'Stock': stock_index, 'ScoreDelta': None} for stock_index in stocks_enabled],
                }
            ]
        }

        # Check number of stocks against min stocks
        if len(stocks_enabled) <= min_stock_count:
            state['Complete'] = True

        return state

    @staticmethod
    def get_enabled_stocks_from_state(state, apply_market_cap_filter=True):
        stocks_enabled = [trial['Stock'] for trial in state['Stages'][-1]['Trials']]

        if state['Stages'][-1]['StockRemoved'] is not None:
            stocks_enabled.remove(state['Stages'][-1]['StockRemoved'])
        
        if apply_market_cap_filter:
            stocks_enabled = list(set(stocks_enabled) - set(state['MarketCapFiltered']))

        return sorted(stocks_enabled)


    def __init__(
        self, 
        returns_data, 
        factors_data, 
        st_duration,
        st_weight,
        objective,
        score_pairs_range,
        min_stock_count,
        score_impact_tolerance,
        state,
        ):
        self.returns_data = returns_data
        self.factors_data = factors_data

        assert(st_duration >= 0)
        assert(st_weight >= 0.0)
        self.st_weight = st_weight

        self.st_months = None
        self.st_period_start = None
        if (self.st_weight > 0.0) and (st_duration > 0):
            self.st_months = self.returns_data.columns.tolist()[-st_duration:]
            self.st_period_start = self.st_months[0]

        self.objective = objective
        
        self.score_pairs_range = score_pairs_range

        self.min_stock_count = min_stock_count
        self.score_impact_tolerance = score_impact_tolerance

        self.state = state

        self.combined_factors = None

    def run(self):

        if self.state['Complete']:
            return

        # Calculate next score
        self._calc_next_score()

        # Update state
        self._update_state()

    def _update_state(self):

        # Check for a completed stage
        stage = self.state['Stages'][-1]

        if stage['BaselineScore'] is None:
            return

        max_score_delta = -np.inf
        max_score_delta_stock = None
        for trial in stage['Trials']:
            if trial['ScoreDelta'] is None:
                return
            if trial['ScoreDelta'] > max_score_delta:
                max_score_delta = trial['ScoreDelta']
                max_score_delta_stock = trial['Stock']

        assert(max_score_delta_stock is not None)

        # Stage is complete - check score
        if max_score_delta > self.score_impact_tolerance:
            stage['StockRemoved'] = max_score_delta_stock
        else:
            self.state['Complete'] = True
            return

        # Check number of stocks
        if len(stage['Trials']) - 1 <= self.min_stock_count:
            self.state['Complete'] = True
            return

        # Add new stage
        stocks = [trial['Stock'] for trial in stage['Trials'] if trial['Stock'] != stage['StockRemoved']]
        self.state['Stages'].append(
            {
                'StockRemoved': None,
                'BaselineScore': max_score_delta + stage['BaselineScore'],
                'Trials': [{'Stock': stock_index, 'ScoreDelta': None} for stock_index in stocks],
            }
        )

    def _calc_next_score(self):

        # Combine factors if needed
        if  self.combined_factors is None:
             self.combined_factors = factor_calc.combine_factors(self.factors_data, list(self.factors_data.keys()))

        # Construct list of stocks
        stage = self.state['Stages'][-1]

        stocks = set([trial['Stock'] for trial in stage['Trials']])
        trial_stock = None
        if stage['BaselineScore'] is not None:
            for trial in stage['Trials']:
                if trial['ScoreDelta'] is None:
                    trial_stock = trial['Stock']
                    stocks.remove(trial_stock)
                    break
        stocks = list(stocks)

        # Apply list of stocks to combined factors and returns
        returns_df = self.returns_data.loc[stocks]
        factor_av = self.combined_factors.loc[stocks]

        # Calculate score

        if self.objective == 'score':

            deltas = factor_calc.calc_long_short_pair_return_deltas(factor_av, returns_df, self.score_pairs_range[0], self.score_pairs_range[1])
            metrics = factor_calc.calc_long_short_pairs_metrics(deltas)
            score = score_calc.calc_score(metrics)

            if self.st_weight > 0.0:
                st_metrics = factor_calc.calc_long_short_pairs_metrics(deltas, start_date=self.st_period_start)
                st_score = score_calc.calc_score(st_metrics)
                score = (score + self.st_weight*st_score)/(1.0 + self.st_weight)

        elif (self.objective == 'corr') or (self.objective == 'rank_corr'):

            corr_method = 'spearman' if self.objective == 'rank_corr' else 'pearson'

            corrs = factor_av.corrwith(returns_df, axis=0, method=corr_method)
            score = corrs.mean()

            if self.st_months is not None:
                st_score = corrs.loc[self.st_months].mean()
                score = (score + self.st_weight*st_score)/(1.0 + self.st_weight)

            score *= 100.0

        else:
            raise ValueError("Invalid objective method '{}'".format(self.objective))

        # Store score
        if stage['BaselineScore'] is None:
            stage['BaselineScore'] = score
        else:
            for trial in stage['Trials']:
                if trial['ScoreDelta'] is None:
                    trial['ScoreDelta'] = score - stage['BaselineScore']
                    break


class StockSelectionReportGenerator():
    
    def __init__(self, db, project_id):
        
        self.db = db
        self.project_id = project_id

        self.bold_font = openpyxl.styles.Font(b=True)
        
    def generate(self):

        self.data_info = self.db.get_project_data_info(self.project_id)

        self.stock_info_df = pd.DataFrame(self.data_info['Stocks'])[['Name','Ticker','SubSector']]
        self.stock_info_df = self.stock_info_df.rename(columns={'SubSector': 'Sub-Sector'})

        self.stocks_disabled = get_indexes_for_disabled_stocks(
            self.data_info['Stocks'], self.db.get_project_stocks_disabled(self.project_id))

        data = self.db.get_project_data(self.project_id, self.data_info, [self.data_info['MarketCapFactorIndex']])
        self.market_cap = data['Factors'][self.data_info['MarketCapFactorIndex']]

        ss_list = self.db.get_project_stock_selection_state(self.project_id)
        ss_list = sorted(ss_list, key=lambda x: x['TargetMonth'])
        
        self.target_months = []
        self.ss_df_list = []
        self.impact_summary_df = self.stock_info_df.copy()
        self.impact_summary_df['Sum'] = 0.0
        self.selection_summary_df = self.stock_info_df.copy()

        for ss in ss_list:
            self._process_target_month(ss)

        self.subsector_impact_summary_df = self.impact_summary_df.groupby('Sub-Sector')['Sum'].mean().to_frame().reset_index()

        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)

        ws = wb.create_sheet('Removal Summary')
        self._write_dataframe(ws, self.selection_summary_df)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25

        ws = wb.create_sheet('Step 1 Impact')
        self._write_dataframe(ws, self.impact_summary_df)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25

        ws = wb.create_sheet('Sub-Sector Impact')
        self._write_dataframe(ws, self.subsector_impact_summary_df)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 17

        for index, ss_df in enumerate(self.ss_df_list):
            ws = wb.create_sheet("{:%b-%Y}".format(self.target_months[index]))
            self._write_dataframe(ws, ss_df)
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25

        return wb
    
    def _write_dataframe(self, ws, df):
        
        for r in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for col in range(1, len(df.columns)+1):
            ws.cell(row=1, column=col).font = self.bold_font
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = 15

        ws.auto_filter.ref = ws.dimensions

    def _process_target_month(self, ss):

        target_month = ss['TargetMonth']
        self.target_months.append(target_month)
        
        target_month_str = "{:%b-%Y}".format(target_month)
        
        df = self.stock_info_df.copy()

        df['Enabled'] = True
        if self.stocks_disabled is not None:
            df.loc[self.stocks_disabled, 'Enabled'] = False

        df["Market Cap".format(target_month_str)] = self.market_cap.loc[target_month]
        
        baseline_score = ss['Stages'][0].get('BaselineScore')
        if baseline_score is not None:
            df['Baseline Score'] = baseline_score

        if (len(ss['Stages']) == 1) and all([trial['ScoreDelta'] is None for trial in ss['Stages'][0]['Trials']]):
            # Nothing removed by score impact
            if ss['Complete']:
                selected = [trial['Stock'] for trial in ss['Stages'][0]['Trials']]
                df['Selected'] = False
                df.loc[selected, 'Selected'] = True
        else:
            for stage_index, stage in enumerate(ss['Stages']):
                score_delta = pd.DataFrame(stage['Trials']).set_index('Stock').astype(float)
                df['Removal Impact (Step {})'.format(stage_index+1)] = score_delta
                if stage_index == 0:
                    col_name = target_month_str
                    self.impact_summary_df[col_name] = score_delta
                    self.impact_summary_df['Sum'] += self.impact_summary_df[col_name]

            if ss['Complete']:
                df['Selected'] = df[df.columns[-1]].notnull()
                last_removed = ss['Stages'][-1]['StockRemoved']
                if last_removed is not None:
                    df.loc[last_removed, 'Selected'] = False

        self.ss_df_list.append(df)

        if 'Selected' in df.columns:
            if 'Removal Impact (Step 1)' in df.columns:
                self.selection_summary_df[target_month_str] = "Impact"
                self.selection_summary_df.loc[df['Removal Impact (Step 1)'].isnull(), target_month_str] = "MC Filter"
            else:
                self.selection_summary_df[target_month_str] = "MC Filter"
            self.selection_summary_df.loc[df['Enabled'] == False, target_month_str] = "Disabled"
            self.selection_summary_df.loc[df['Selected'] == True, target_month_str] = ""

            


    








