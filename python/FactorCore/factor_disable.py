import pandas as pd
import numpy as np
import logging
import openpyxl, openpyxl.utils.dataframe

logger = logging.getLogger(__name__)

class FactorDisabler():

    @staticmethod
    def create_state(factors_enabled):

        assert(factors_enabled is not None)

        state = {
            'Complete': False,
            'UserEnabledFactors': factors_enabled,
            'FactorDataComplete': [],
        }

        return state

    @staticmethod
    def get_enabled_factors_from_state(state):
        enabled = [item['Factor'] for item in state['FactorDataComplete'] if item['Enabled']]
        return enabled

    def __init__(
        self, 
        factors_data,
        date_range,
        required_data_fraction,
        stocks_enabled,
        state,
        ):
        self.factors_data = factors_data
        self.date_range = date_range
        self.required_data_fraction = required_data_fraction
        self.stocks_enabled = stocks_enabled
        self.state = state

    def run(self):

        if self.state['Complete']:
            return

        self.state['FactorDataComplete'] = []

        for factor in self.state['UserEnabledFactors']:

            df = self.factors_data[factor]

            df = df.loc[self.date_range[0]:self.date_range[1]][self.stocks_enabled]

            stock_count = len(df.columns)

            factor_count = df.count(axis=1)

            min_factor_fraction = float((factor_count / stock_count).min())

            self.state['FactorDataComplete'].append({'Factor': factor, 'MinDataFraction': min_factor_fraction, 'Enabled': min_factor_fraction > self.required_data_fraction})


class FactorsDisabledReportGenerator():
    
    def __init__(self, db, project_id):
        
        self.db = db
        self.project_id = project_id

        self.bold_font = openpyxl.styles.Font(b=True)
        
    def generate(self):

        data_info = self.db.get_project_data_info(self.project_id)

        info_df = pd.DataFrame(data_info['Factors'], columns=['Factor'])

        state_list = self.db.get_project_factor_disabled_state(self.project_id)

        min_data_fractions = {}
        enabled = {}
        for state in state_list:
            month_min_data_fractions = {}
            month_enabled = {}
            for item in state['FactorDataComplete']:
                month_min_data_fractions[item['Factor']] = item['MinDataFraction']
                month_enabled[item['Factor']] = item['Enabled']
            month = "{:%b-%Y}".format(state['TargetMonth'])
            min_data_fractions[month] = month_min_data_fractions
            enabled[month] = month_enabled

        min_data_fractions = pd.concat([info_df, pd.DataFrame(min_data_fractions)], axis=1)

        disabled = pd.concat([info_df, pd.DataFrame(enabled)], axis=1).set_index('Factor').fillna(False)
        disabled = (disabled == False).replace(False, "").replace(True, "Disabled").reset_index()

        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)

        ws = wb.create_sheet('Disabled')
        self._write_dataframe(ws, disabled)

        ws = wb.create_sheet('Data Completeness')
        self._write_dataframe(ws, min_data_fractions)

        return wb
    
    def _write_dataframe(self, ws, df):
        
        for r in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for col in range(1, len(df.columns)+1):
            ws.cell(row=1, column=col).font = self.bold_font
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = 15

        ws.auto_filter.ref = ws.dimensions
        
            
