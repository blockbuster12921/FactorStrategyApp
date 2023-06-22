import numpy as np
import pandas as pd
import openpyxl
import logging
from . import settings

logger = logging.getLogger(__name__)

def combine_stock_rankings(ranked_stocks_list):
    
    dfs = []
    for ranked_stocks in ranked_stocks_list:
        df = pd.Series(index=ranked_stocks, data=np.array(range(0, len(ranked_stocks)))/len(ranked_stocks)).sort_index()
        dfs.append(df)
        
    return pd.concat(dfs, axis=1).mean(axis=1).sort_values().index.tolist()

def calc_long_short_returns_from_ranks(ranks, returns, min_pairs, max_pairs):

    assert(isinstance(returns, pd.Series))

    r = returns.loc[ranks].dropna().values

    deltas = {}
    delta_sum = 0.0
    for i in range(0, min(max_pairs, int(len(r)/2))):
        delta = r[i] - r[-(i+1)]
        delta_sum += delta
        if i+1 >= min_pairs:
            deltas[i+1] = delta_sum / (i+1)

    return deltas


class StockRankingReportGenerator():
    
    def __init__(self, db, strategy_id):
        self.db = db
        self.strategy_id = strategy_id
        
        self.bold_font = openpyxl.styles.Font(b=True)
        self.left_alignment = openpyxl.styles.Alignment(horizontal='left')
        self.centre_alignment = openpyxl.styles.Alignment(horizontal='center')

    def generate_for_project(self, project_id):

        # Load data and settings
        project = self.db.get_project(project_id)

        data_info = self.db.get_project_data_info(project_id)

        results = self.db.get_project_factor_strategy_results(project_id, strategy_id=self.strategy_id, contexts=None)
        if len(results) == 0:
            return None

        project_settings = self.db.get_project_settings(project_id)
        project_settings = settings.overlay_default_project_settings(project_settings)
        strategy_ids = project_settings['FactorSelectionStrategies'] if self.strategy_id is None else [self.strategy_id]
        strategy_indexes = { strategy_id: index for index, strategy_id in enumerate(strategy_ids) }

        data = self.db.get_project_data(project_id, data_info, factor_indexes=[])
        self.returns_df = data['Returns']

        # Collate sources and target months
        sources = set()
        target_months = set()
        for result in results:
            target_months.add(result['TargetMonth'])
            sources.add(result['Context'])
        target_months = sorted(target_months)

        # Write Excel report
        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)
        
        for target_month in target_months:

            ws = wb.create_sheet("{:%b-%Y}".format(target_month))
            start_row = 1

            for strategy_id in strategy_ids:
                for source in sources:
                    
                    for result in results:
                        if (result['TargetMonth'] == target_month) and (result['StrategyID'] == strategy_id) and (result.get('Context') == source):
                            end_row = self._generate_single_strategy(project['Name'], target_month, data_info, result, ws, start_row)
                            start_row = end_row + 2
                            break

        return wb

    def generate_for_project_group(self, project_group_id):

        project_group = self.db.get_project_group(project_group_id)

        wb = openpyxl.Workbook(write_only=False)
        wb.remove(wb.active)

        for project_id in project_group['ProjectIDs']:

            # Load data and settings
            project = self.db.get_project(project_id)

            project_settings = self.db.get_project_settings(project_id)
            project_settings = settings.overlay_default_project_settings(project_settings)

            data_info = self.db.get_project_data_info(project_id)

            strategy_id = self.strategy_id
            if strategy_id is None:
                # Get single strategy from settings
                if len(project_settings['FactorSelectionStrategies']) != 1:
                    raise ValueError("Project {} has {} active Factor Selection Strategies - it must have exactly one".format(
                                     project['Name'], len(project_settings['FactorSelectionStrategies'])))
                strategy_id = project_settings['FactorSelectionStrategies'][0]
                print(strategy_id, flush=True)

            results = self.db.get_project_factor_strategy_results(project_id, strategy_id=strategy_id, contexts=None)
            if len(results) == 0:
                continue

            data = self.db.get_project_data(project_id, data_info, factor_indexes=[])
            self.returns_df = data['Returns']

            # Collate sources and target months
            sources = set()
            target_months = set()
            for result in results:
                target_months.add(result['TargetMonth'])
                sources.add(result['Context'])
            target_months = sorted(target_months)

            # Write Excel report
            
            for target_month in target_months:

                if len(target_months) == 1:
                    ws_name = project['Name'][:31]
                else:
                    ws_name = "{} ({:%b-%y})".format(project['Name'][:22], target_month)

                ws = wb.create_sheet(ws_name)
                start_row = 1

                for source in sources:
                    for result in results:
                        if (result['TargetMonth'] == target_month) and (result['StrategyID'] == strategy_id) and (result.get('Context') == source):
                            end_row = self._generate_single_strategy(project['Name'], target_month, data_info, result, ws, start_row)
                            start_row = end_row + 2
                            break

        return wb

    def _generate_single_strategy(self, project_name, target_month, data_info, result, ws, start_row):

        if result.get('RankedStocks') is None:
            # No ranked stocks
            return start_row

        # Info
        ws.cell(row=start_row, column=1, value="Project").font = self.bold_font
        ws.cell(row=start_row, column=2, value=project_name)

        ws.cell(row=start_row+1, column=1, value="Month").font = self.bold_font
        ws.cell(row=start_row+1, column=2, value="{:%b-%Y}".format(target_month))

        ws.cell(row=start_row+3, column=1, value="Combinations").font = self.bold_font
        combinations_text = result['StrategyDescription']['combinations']
        if result['Context'] is not None:
            combinations_text += " (from {})".format(result['Context'].capitalize())
        ws.cell(row=start_row+3, column=2, value=combinations_text)
        ws.cell(row=start_row+4, column=1, value="Factors").font = self.bold_font
        ws.cell(row=start_row+4, column=2, value=result['StrategyDescription']['factors'])
        ws.cell(row=start_row+5, column=1, value="Weighting").font = self.bold_font
        ws.cell(row=start_row+5, column=2, value=result['StrategyDescription']['weighting'])
        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 5
        
        # Stocks
        cell = ws.cell(row=start_row, column=4, value="Rank")
        cell.font = self.bold_font
        cell.alignment = self.centre_alignment
        ws.cell(row=start_row, column=5, value="Long (Name)").font = self.bold_font
        ws.cell(row=start_row, column=6, value="Long (Ticker)").font = self.bold_font
        ws.cell(row=start_row, column=7, value="Short (Name)").font = self.bold_font
        ws.cell(row=start_row, column=8, value="Short (Ticker)").font = self.bold_font
        ws.column_dimensions['D'].width = 7
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15

        returns = self.returns_df.loc[result['TargetMonth']]
        returns_count = returns.count()
        if returns_count > 0:
            ws.cell(row=start_row, column=9, value="Long Return").font = self.bold_font
            ws.cell(row=start_row, column=10, value="Short Return").font = self.bold_font
            ws.cell(row=start_row, column=11, value="Long/Short Return").font = self.bold_font
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 15

        row = start_row
        for rank, long_stock_index in enumerate(result['RankedStocks']):

            if rank >= len(result['RankedStocks'])/2:
                break

            long_stock = data_info['Stocks'][long_stock_index]
            short_stock_index = result['RankedStocks'][len(result['RankedStocks'])-rank-1]
            short_stock = data_info['Stocks'][short_stock_index]

            long_stock_text = "{} ({})".format(long_stock['Name'], long_stock['Ticker'])
            short_stock_text = "{} ({})".format(short_stock['Name'], short_stock['Ticker'])

            row = start_row + rank + 1
            ws.cell(row=row, column=4, value=rank+1)
            ws.cell(row=row, column=5, value=long_stock['Name'])
            ws.cell(row=row, column=6, value=long_stock['Ticker'])
            if short_stock_index != long_stock_index:
                ws.cell(row=row, column=7, value=short_stock['Name'])
                ws.cell(row=row, column=8, value=short_stock['Ticker'])
            
            if returns_count > 0:
                long_stock_return = returns.loc[long_stock_index]
                if not np.isnan(long_stock_return):
                    cell = ws.cell(row=row, column=9, value=long_stock_return)
                    cell.number_format = '0.0%'
                if short_stock_index != long_stock_index:
                    short_stock_return = returns.loc[short_stock_index]
                    if not np.isnan(short_stock_return):
                        cell = ws.cell(row=row, column=10, value=short_stock_return)
                        cell.number_format = '0.0%'
                    long_short_return = long_stock_return - short_stock_return
                    if not np.isnan(long_short_return):
                        cell = ws.cell(row=row, column=11, value=long_short_return)
                        cell.number_format = '0.0%'

        last_row = row
        
        # Factors
        ws.cell(row=start_row+7, column=1, value="Factor").font = self.bold_font
        ws.cell(row=start_row+7, column=2, value="Weight").font = self.bold_font

        for index, factor_index in enumerate(result['Factors']):
            factor = data_info['Factors'][factor_index]
            weight = result['FactorWeights'][index]
            row = start_row + index + 8
            ws.cell(row=row, column=1, value=factor)
            cell = ws.cell(row=row, column=2, value=weight)
            cell.number_format = '0.0%'
            cell.alignment = self.left_alignment

        last_row = max(last_row, start_row + len(result['Factors']) + 5)

        return last_row


def load_stock_ranking_report(path):
    
    def parse_stock_name(s):
        
        assert(s[-1] == ')')

        bracket_pos = s.rfind('(')
        ticker = s[bracket_pos+1:-1].strip()
        name = s[:bracket_pos-1].strip()

        return { 'Ticker': ticker, 'Name': name }

    logger.info("Loading Stock Ranking report '{}'...".format(path.name))

    wb = openpyxl.load_workbook(filename=str(path), read_only=True)
    
    data = []

    for ws in wb.worksheets:
        data_date = (pd.Timestamp(ws.title) + pd.DateOffset(months=0)) + pd.offsets.MonthEnd(0)
        
        all_cells = [[cell.value for cell in row] for row in ws.iter_rows()]

        start_row = None
        for row in range(0, len(all_cells)):
            
            if all_cells[row][0] == 'Project':
                new_format = True

            if all_cells[row][1] == 'All (from Optimize)':
                start_row = row

                if row >= 3:
                    if all_cells[row-3][0] == 'Project':
                        # New format
                        start_row = row - 3
                    
                break

        if start_row is None:
            raise ValueError("'All (from Optimize)' not found in worksheet '{}'".format(ws.title))
        
        longs = []
        shorts = []
        long_returns = []
        short_returns = []
        
        if all_cells[start_row][4] == 'Long':
            
            # Old format with Name and Ticker in same column

            for row_index in range(start_row+1, len(all_cells)):
                row = all_cells[row_index]
                long_name = row[4]
                short_name = row[5]
                if (long_name is None) or (short_name is None):
                    break
                longs.append(long_name)
                shorts.append(short_name)
                long_returns.append(row[6])
                short_returns.append(row[7])

            assert(len(longs) == len(shorts))

            longs = [parse_stock_name(s) for s in longs]
            shorts = [parse_stock_name(s) for s in shorts]

        else:

            for row_index in range(start_row+1, len(all_cells)):
                row = all_cells[row_index]
                long_name = row[4]
                short_name = row[6]
                if (long_name is None) or (short_name is None):
                    break
                longs.append({'Name': long_name, 'Ticker': row[5]})
                shorts.append({'Name': short_name, 'Ticker': row[7]})
                long_returns.append(row[8])
                short_returns.append(row[9])

            assert(len(longs) == len(shorts))

        factor_weights = {}
        for row_index in range(start_row+5, len(all_cells)):
            row = all_cells[row_index]
            
            factor = row[0]
            weight = row[1]

            if (factor is None) or (weight is None):
                break

            factor_weights[factor] = weight

        pair_returns = pd.Series(index=range(1,1+len(long_returns)), data=long_returns)
        pair_returns -= np.array(short_returns)
        pair_returns.index.name = 'Pairs'
        
        data.append({ 'DataDate': data_date, 'Longs': longs, 'Shorts': shorts,
                      'LongReturns': long_returns, 'ShortReturns': short_returns,
                      'PairReturns': pair_returns, 'FactorWeights': factor_weights })

    return data    

