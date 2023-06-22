import numpy as np
import pandas as pd
import logging
import openpyxl
import copy
import numbers
from . import factor_calc, score_calc, settings, factor_strategy
from .trading_solution import calc_sharpe

logger = logging.getLogger(__name__)

def calc_bootstrap_samples(data, stats, n_samples):
    
    samples = { key: [] for key in stats.keys() }
        
    for i in range(n_samples):
        sample = np.random.choice(data, len(data), replace=True)

        for key, fn in stats.items():
            samples[key].append(fn(sample))

    return pd.DataFrame(samples)

def calc_bootstrap_confidence_interval(data, stats, n_samples=1000, confidence=0.95):

    samples = calc_bootstrap_samples(data, stats, n_samples)

    alpha = 1.0 - confidence
    ci = samples.quantile([alpha/2.0, 1.0-alpha/2.0])

    return ci

def generate_excel_report(results, factor_names, pairs_range, project_settings):

    wb = openpyxl.Workbook(write_only=False)
    wb.remove(wb.active)

    # Settings sheet
    wb.create_sheet('Settings')
    settings_sheet = wb['Settings']
    try:
        project_settings.pop('_id')
        project_settings.pop('FactorSelectionStrategies')
    except:
        pass
    for row, key in enumerate(sorted(project_settings.keys()), 1):
        settings_sheet.cell(row=row, column=1, value=key)
        setting = project_settings[key]
        if not isinstance(setting, numbers.Number):
            setting = str(setting)
        settings_sheet.cell(row=row, column=2, value=setting)
    settings_sheet.column_dimensions['A'].width = 50

    bold_font = openpyxl.styles.Font(b=True)
    blue_font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
    yellow_fill = openpyxl.styles.PatternFill(start_color="FEFFCF", end_color="FEFFCF", fill_type= "solid")

    def generate_summary_panel(ws, results, col, last_data_row, score_col):

        # Summary panel info
        strategy_combinations = []
        strategy_factors = []
        strategy_weightings = []
        for result in results:
            description = result['StrategyDescription']
            if description['combinations'] not in strategy_combinations:
                strategy_combinations.append(description['combinations'])
            if description['factors'] not in strategy_factors:
                strategy_factors.append(description['factors'])
            if description['weighting'] not in strategy_weightings:
                strategy_weightings.append(description['weighting'])

        col_letter = openpyxl.utils.cell.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

        for weighting_col, weighting in enumerate(strategy_weightings, col+1):
            ws.cell(row=1, column=weighting_col, value=weighting).font = bold_font

        score_col_letter = openpyxl.utils.cell.get_column_letter(score_col)

        ws.cell(row=2, column=col, value="Combinations:").font = bold_font
        for row, combination in enumerate(strategy_combinations, 3):
            ws.cell(row=row, column=col, value=combination)
            for weighting_col, weighting in enumerate(strategy_weightings, col+1):
                weighting_col_letter = openpyxl.utils.cell.get_column_letter(weighting_col)
                formula = "=AVERAGEIFS(${}$2:${}${},$B$2:$B${},${}{},$D$2:$D${},${}$1)".format(
                    score_col_letter, score_col_letter, last_data_row, last_data_row, col_letter, row, last_data_row, weighting_col_letter
                )
                cell = ws.cell(row=row, column=weighting_col, value=formula)
                cell.number_format = '0.00'

        chart = openpyxl.chart.BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Combinations"
        chart.title.overlay = True
        chart.y_axis.title = "Score"
        for weighting_col, weighting in enumerate(strategy_weightings, col+1):
            chart_data = openpyxl.chart.Reference(ws, min_col=weighting_col, min_row=3, max_row=row, max_col=weighting_col)
            chart.append(openpyxl.chart.Series(chart_data, title=weighting))
        chart_cats = openpyxl.chart.Reference(ws, min_col=col, max_col=col, min_row=3, max_row=row)
        chart.set_categories(chart_cats)
        chart.height = 10
        chart.width = 30
        chart.legend.position = 'b'
        ws.add_chart(chart, "{}{}".format(openpyxl.utils.cell.get_column_letter(col+len(strategy_weightings)+2), 2))

        row += 2
        start_row = row
        ws.cell(row=row, column=col, value="Factors:").font = bold_font
        factors_rows = {}
        for row, factor in enumerate(strategy_factors, row+1):
            ws.cell(row=row, column=col, value=factor)
            factors_rows[factor] = row
            for weighting_col, weighting in enumerate(strategy_weightings, col+1):
                weighting_col_letter = openpyxl.utils.cell.get_column_letter(weighting_col)
                formula = "=AVERAGEIFS(${}$2:${}${},$C$2:$C${},${}{},$D$2:$D${},${}$1)".format(
                    score_col_letter, score_col_letter, last_data_row, last_data_row, col_letter, row, last_data_row, weighting_col_letter
                )
                cell = ws.cell(row=row, column=weighting_col, value=formula)
                cell.number_format = '0.00'

        chart = openpyxl.chart.BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Factors"
        chart.title.overlay = True
        chart.y_axis.title = "Score"
        for weighting_col, weighting in enumerate(strategy_weightings, col+1):
            chart_data = openpyxl.chart.Reference(ws, min_col=weighting_col, min_row=start_row+1, max_row=row, max_col=weighting_col)
            chart.append(openpyxl.chart.Series(chart_data, title=weighting))
        chart_cats = openpyxl.chart.Reference(ws, min_col=col, max_col=col, min_row=start_row+1, max_row=row)
        chart.set_categories(chart_cats)
        chart.height = 10
        chart.width = 30
        chart.legend.position = 'b'
        ws.add_chart(chart, "{}{}".format(openpyxl.utils.cell.get_column_letter(col+len(strategy_weightings)+2), max(22, start_row)))

        row += 2
        ws.cell(row=row, column=col, value="Factor & Combinations:").font = bold_font
        row += 1
        ws.cell(row=row, column=col, value=None if len(strategy_factors) == 0 else strategy_factors[0]).font = blue_font
        factor_choice_row = row
        for row, combination in enumerate(strategy_combinations, row+1):
            ws.cell(row=row, column=col, value=combination)
            for weighting_col, weighting in enumerate(strategy_weightings, col+1):
                weighting_col_letter = openpyxl.utils.cell.get_column_letter(weighting_col)
                formula = "AVERAGEIFS(${}$2:${}${},$B$2:$B${},${}{},$D$2:$D${},${}$1,$C$2:$C${},${}${})".format(
                    score_col_letter, score_col_letter, last_data_row, last_data_row, col_letter, row, last_data_row, weighting_col_letter, last_data_row, col_letter, factor_choice_row
                )
                formula = '=IF(ISNUMBER({}),{},"")'.format(formula, formula)
                cell = ws.cell(row=row, column=weighting_col, value=formula)
                cell.number_format = '0.00'

        row += 2
        ws.cell(row=row, column=col, value="Ex:").font = bold_font
        row += 1
        for factor in strategy_factors:
            ex_pos = factor.find(' ex ')
            if ex_pos <= 0:
                continue
            factor_root_row = factors_rows.get(factor[:ex_pos])
            if factor_root_row is None:
                continue
            factor_row = factors_rows.get(factor)

            ws.cell(row=row, column=col, value=factor)
            for weighting_col, weighting in enumerate(strategy_weightings, col+1):
                weighting_col_letter = openpyxl.utils.cell.get_column_letter(weighting_col)
                formula = "${}${}-${}${}".format(
                    weighting_col_letter, factor_row, weighting_col_letter, factor_root_row
                )
                formula = '=IF(ISNUMBER({}),{},"")'.format(formula, formula)
                cell = ws.cell(row=row, column=weighting_col, value=formula)
                cell.number_format = '0.00'
            row += 1

        for panel_col in range(col, col+3):
            for panel_row in range(1, row):
                ws.cell(row=panel_row, column=panel_col).fill = yellow_fill

        summary_range = { 'StartRow': 1, 'EndRow': row, 'StartCol': col, 'EndCol': col+2, 'Sheet': ws.title }
        return summary_range 


    def generate_pairs_summary_panel(ws, pairs_range, col, last_data_row, score_col):

        col_letter = openpyxl.utils.cell.get_column_letter(col)

        row = 3
        cell = ws.cell(row=row, column=col, value=ws.cell(row=2, column=3).value)
        cell.fill = yellow_fill
        cell.font = blue_font
        cell = ws.cell(row=row, column=col+1, value=ws.cell(row=2, column=4).value)
        cell.font = blue_font
        cell.fill = yellow_fill

        score_col_letter = openpyxl.utils.cell.get_column_letter(score_col)
        formula_col_letter = openpyxl.utils.cell.get_column_letter(col+1)

        start_row = row
        for row, pairs in enumerate(range(pairs_range[0], pairs_range[1]+1), start_row+1):
            cell = ws.cell(row=row, column=col, value=pairs)
            cell.fill = yellow_fill
            formula = "=AVERAGEIFS(${}$2:${}${},$C$2:$C${},${}${},$D$2:$D${},${}$3,$E$2:$E${},{}{})".format(
                    score_col_letter, score_col_letter, last_data_row, last_data_row, col_letter, start_row, last_data_row, formula_col_letter, last_data_row, col_letter, row
                )
            cell = ws.cell(row=row, column=col+1, value=formula)
            cell.number_format = '0.00'
            cell.fill = yellow_fill

        chart = openpyxl.chart.BarChart()
        chart.type = "col"
        chart.style = 10
        chart.y_axis.title = "Score"
        chart_data = openpyxl.chart.Reference(ws, min_col=col+1, max_col=col+1, min_row=start_row+1, max_row=row)
        chart.add_data(chart_data, titles_from_data=False)
        chart_cats = openpyxl.chart.Reference(ws, min_col=col, max_col=col, min_row=start_row+1, max_row=row)
        chart.set_categories(chart_cats)
        chart.legend = None
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, "{}{}".format(openpyxl.utils.cell.get_column_letter(col+3), 3))



    def generate_scores_sheet(ws, results, pairs_range):

        row = ['Strategy','Combinations','Factors','Weighting']
        row += ["{} - {} pairs".format(pairs_range[0], pairs_range[1])]
        row += list(range(pairs_range[0], pairs_range[1]+1))
        ws.append(row)

        for strategy_number, result in enumerate(results, 1):
            description = result['StrategyDescription']
            row = [strategy_number]
            row += [description[key] for key in ['combinations','factors','weighting']]
            row += [result['OverallScore']]
            row += [result['PairScores'].get(pairs) for pairs in range(pairs_range[0], pairs_range[1]+1)]
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=1+len(results), min_col=5, max_col=6+pairs_range[1]-pairs_range[0]):
            for cell in row:
                cell.number_format = '0.00'

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14

        header_font = openpyxl.styles.Font(b=True)
        header_align = openpyxl.styles.Alignment(horizontal='left')
        for cell in ws['1:1']:
            cell.font = header_font
            cell.alignment = header_align

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'E2'

        return generate_summary_panel(ws, results, 8+pairs_range[1]-pairs_range[0], len(results)+1, score_col=5)


    def generate_returns_detail(ws, results, pairs_range):

        if len(results) == 0:
            return
        
        row = ['Strategy','Combinations','Factors','Weighting','Pairs','Score','Sharpe','Sharpe 95% CI Low','Sharpe 95% CI High','Mean','Mean 95% CI Low','Mean 95% CI High','Std Dev']
        ws.append(row)
        returns_col_start = 1+len(row)

        for col, month_result in enumerate(results[0]['Months'], returns_col_start):
            cell = ws.cell(row=1, column=col)
            cell.value = pd.Timestamp(month_result['TargetMonth'])
            cell.number_format = 'mmm-yy'

        row = 2
        for result_index, result in enumerate(results):

            description = result['StrategyDescription']

            for pairs in range(pairs_range[0], pairs_range[1]+1):

                pairs_row = row + pairs - pairs_range[0]
                
                ws.cell(row=pairs_row, column=1, value=result_index+1)
                ws.cell(row=pairs_row, column=2, value=description['combinations'])
                ws.cell(row=pairs_row, column=3, value=description['factors'])
                ws.cell(row=pairs_row, column=4, value=description['weighting'])
                ws.cell(row=pairs_row, column=5, value=pairs)

                score = result['PairScores'].get(pairs)
                if score is not None:
                    cell = ws.cell(row=pairs_row, column=6)
                    cell.value = score
                    cell.number_format = '0.00'

                sharpe = result['Sharpes'].get(pairs)
                if sharpe is not None:
                    cell = ws.cell(row=pairs_row, column=7)
                    cell.value = sharpe
                    cell.number_format = '0.00'

                    ci = result.get('SharpeCIs')
                    if ci is not None:
                        ci = ci.get(pairs)
                        if ci is not None:
                            cell = ws.cell(row=pairs_row, column=8)
                            cell.value = ci[0]
                            cell.number_format = '0.00'
                            cell = ws.cell(row=pairs_row, column=9)
                            cell.value = ci[1]
                            cell.number_format = '0.00'

                mean = result['MeanReturns'].get(pairs)
                if mean is not None:
                    cell = ws.cell(row=pairs_row, column=10)
                    cell.value = mean
                    cell.number_format = '0.00%'

                    ci = result['MeanReturnCIs']
                    if ci is not None:
                        ci = ci.get(pairs)
                        if ci is not None:
                            cell = ws.cell(row=pairs_row, column=11)
                            cell.value = ci[0]
                            cell.number_format = '0.00%'
                            cell = ws.cell(row=pairs_row, column=12)
                            cell.value = ci[1]
                            cell.number_format = '0.00%'

                std_dev = result['ReturnStdDevs'].get(pairs)
                if std_dev is not None:
                    cell = ws.cell(row=pairs_row, column=13)
                    cell.value = std_dev
                    cell.number_format = '0.00%'

            for col, month_result in enumerate(result['Months'], returns_col_start):
                for pairs in range(pairs_range[0], pairs_range[1]+1):
                    return_value = None if month_result['ReturnDeltas'] is None else month_result['ReturnDeltas'].get(pairs)
                    if return_value is not None:
                        cell = ws.cell(row=row+pairs-pairs_range[0], column=col)
                        cell.value = return_value
                        cell.number_format = '0.00%'

            row += 1 + pairs_range[1] - pairs_range[0]
            
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 10

        header_font = openpyxl.styles.Font(b=True)
        header_align = openpyxl.styles.Alignment(horizontal='left')
        for cell in ws['1:1']:
            cell.font = header_font
            cell.alignment = header_align

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'N2'

        if pairs_range[0] == pairs_range[1]:
            return generate_summary_panel(ws, results, 1 + returns_col_start + len(results[0]['Months']), len(results)+1, score_col=6)
        else:
            generate_pairs_summary_panel(ws, pairs_range, 1 + returns_col_start + len(results[0]['Months']), (1+pairs_range[1]-pairs_range[0])*len(results)+1, score_col=6)

    def generate_correlations(ws, results):

        if len(results) == 0:
            return
        
        row = ['Strategy','Combinations','Factors','Weighting','Mean','Mean 95% CI Low','Mean 95% CI High']
        ws.append(row)
        months_col_start = 1+len(row)

        for col, month_result in enumerate(results[0]['Months'], months_col_start):
            cell = ws.cell(row=1, column=col)
            cell.value = pd.Timestamp(month_result['TargetMonth'])
            cell.number_format = 'mmm-yy'

        row = 2
        for result_index, result in enumerate(results):

            description = result['StrategyDescription']

            ws.cell(row=row, column=1, value=result_index+1)
            ws.cell(row=row, column=2, value=description['combinations'])
            ws.cell(row=row, column=3, value=description['factors'])
            ws.cell(row=row, column=4, value=description['weighting'])

            mean_corr = result['MeanRankCorrelation']
            if mean_corr is not None:
                cell = ws.cell(row=row, column=5)
                cell.value = mean_corr
                cell.number_format = '0.0%'

            ci = result.get('MeanRankCorrelationCI')
            if ci is not None:
                cell = ws.cell(row=row, column=6)
                cell.value = ci[0]
                cell.number_format = '0.0%'
                cell = ws.cell(row=row, column=7)
                cell.value = ci[1]
                cell.number_format = '0.0%'

            for col, month_result in enumerate(result['Months'], months_col_start):

                corr = month_result['RankCorrelation']
                if corr is not None:
                    cell = ws.cell(row=row, column=col)
                    cell.value = corr
                    cell.number_format = '0.0%'

            row += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

        header_font = openpyxl.styles.Font(b=True)
        header_align = openpyxl.styles.Alignment(horizontal='left')
        for cell in ws['1:1']:
            cell.font = header_font
            cell.alignment = header_align

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'H2'


    def generate_factors_detail(ws, results, pairs_range):
        
        if len(results) == 0:
            return
        
        row = ['Combinations','Factors','Weighting','Month']
        factors_col_start = 1+len(row)
        for f in factor_names:
            row.append(f)
            row.append("-" + f)
        ws.append(row)

        row = 2
        for result in results:

            description = result['StrategyDescription']
            
            for month_index, month_result in enumerate(result['Months'], 0):

                ws.cell(row=row, column=1, value=description['combinations'])
                ws.cell(row=row, column=2, value=description['factors'])
                ws.cell(row=row, column=3, value=description['weighting'])

                cell = ws.cell(row=row, column=4)
                cell.value = pd.Timestamp(month_result['TargetMonth'])
                cell.number_format = 'mmm-yy'

                factors = month_result['Factors']
                weights = month_result['FactorWeights']

                if (factors is not None) and (weights is not None):
                    weight_sum = sum(weights)
                    for f_index, f in enumerate(factors):
                        f_pos = 2*f if f >= 0 else 2*(-f-1)+1
                        cell = ws.cell(row=row, column=factors_col_start+f_pos)
                        cell.value = weights[f_index]/weight_sum
                        cell.number_format = '0.00%'

                row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12

        header_font = openpyxl.styles.Font(b=True)
        header_align = openpyxl.styles.Alignment(horizontal='left')
        for cell in ws['1:1']:
            cell.font = header_font
            cell.alignment = header_align

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'E2'

    def generate_summary_range(ws, source_ws, summary_range, last_col=-1):

        row_count = summary_range['EndRow'] - summary_range['StartRow'] + 1
        col_count = summary_range['EndCol'] - summary_range['StartCol'] + 1

        start_col = last_col + 2
        start_row = 2

        ws.cell(row=start_row-1, column=start_col, value=source_ws.title).font = bold_font

        for col in range(start_col, start_col+col_count):
            source_col = summary_range['StartCol'] + col - start_col 
            source_col_letter = openpyxl.utils.cell.get_column_letter(source_col)
            col_width = 20 if col == start_col else 12
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = col_width
            for row in range(start_row, row_count+2):
                source_row = summary_range['StartRow'] + row - start_row
                formula = "'{}'!${}${}".format(source_ws.title, source_col_letter, source_row)
                formula = '=IF(ISBLANK({}),"",{})'.format(formula, formula)
                cell = ws.cell(row=row, column=col, value=formula)
                source_cell = source_ws.cell(row=source_row, column=source_col)
                cell.font = copy.copy(source_cell.font)
                cell.fill = copy.copy(source_cell.fill)
                cell.number_format = copy.copy(source_cell.number_format)

        return start_col + col_count - 1


    summary_range_source_sheets = {}
    summary_sheet = wb.create_sheet('Summary')

    source_list = ['optimize','generate']

    for source in source_list:

        source_results = results.get(source)
        if source_results is None:
            continue

        source_results = [res for res in source_results if res is not None]
        if len(source_results) == 0:
            continue

        ws_prefix = '' if source is None else source.capitalize()[:3] + ' '

        ws = wb.create_sheet('{}Scores'.format(ws_prefix))
        summary_range = generate_scores_sheet(ws, source_results, pairs_range)
        summary_range_source_sheets[ws.title] = summary_range

        ws = wb.create_sheet('{}Returns'.format(ws_prefix))
        generate_returns_detail(ws, source_results, pairs_range)

        target_pairs = int((pairs_range[0] + pairs_range[1])/2)
        for pairs in [target_pairs-1, target_pairs, target_pairs+1]:
            ws = wb.create_sheet('{}Returns ({}s)'.format(ws_prefix, pairs))
            summary_range = generate_returns_detail(ws, source_results, (pairs,pairs))
            summary_range_source_sheets[ws.title] = summary_range

        ws = wb.create_sheet('{}Correlations'.format(ws_prefix))
        generate_correlations(ws, source_results)

        ws = wb.create_sheet('{}Factors'.format(ws_prefix))
        generate_factors_detail(ws, source_results, pairs_range)

    last_col = -1
    for source_sheet in summary_range_source_sheets:

        summary_range = summary_range_source_sheets.get(source_sheet)
        if summary_range is None:
            continue

        last_col = generate_summary_range(summary_sheet, wb[source_sheet], summary_range, last_col)

    return wb

def collate_factor_strategy_results(db, project_id):

    project_settings = db.get_project_settings(project_id)
    project_settings = settings.overlay_default_project_settings(project_settings)
    strategy_ids = project_settings['FactorSelectionStrategies']
    strategies = db.get_factor_strategies(strategy_ids)
    strategy_indexes = { strategy['ID']: index for index, strategy in enumerate(strategies) }

    results = db.get_project_factor_strategy_results(project_id, strategy_id=None, contexts=None)

    project_data_info = db.get_project_data_info(project_id)
    returns_df = db.get_project_data(project_id, project_data_info, factor_indexes=[])
    if returns_df is not None:
        returns_df = returns_df['Returns']

    # Check that we have results for all enabled strategies and recalculate otherwise
    result_strategy_ids = set([result['StrategyID'] for result in results])
    missing_strategy_ids = list(set(strategy_ids) - set(result_strategy_ids))
    if len(missing_strategy_ids) > 0:
        factor_strategy.calculate_project_factor_strategies(db, project_id, missing_strategy_ids)
        results = db.get_project_factor_strategy_results(project_id, strategy_id=None, contexts=None)
    
    if len(results) == 0:
        return None
    
    sources = set()
    target_months = set()
    for result in results:
        target_months.add(result['TargetMonth'])
        sources.add(result['Context'])
    target_months = sorted(target_months)

    results_by_source = { source: [] for source in sources }
    for source in results_by_source.keys():
        for strategy in strategies:
            results_by_source[source].append(
                { 'StrategyDescription': strategy['Description'], 'Months': [] })

    for result in results:
        month_result = {
            'TargetMonth': result['TargetMonth'],
            'Factors': result['Factors'],
            'FactorWeights': result['FactorWeights'],
            'ReturnDeltas': None,
            }

        if result['ReturnDeltas'] is not None:
            deltas = { pair: result['ReturnDeltas']['Values'][index] for index, pair in enumerate(result['ReturnDeltas']['Pairs']) }
            month_result['ReturnDeltas'] = deltas

        ranked_stocks = result.get('RankedStocks')
        if (returns_df is not None) and (ranked_stocks is not None):
            month_returns = returns_df.loc[result['TargetMonth']]
            if len(month_returns) > 0:
                stock_ranks = pd.Series({ stock: i for i, stock in enumerate(ranked_stocks, 1) }).to_frame('Rank')
                stock_ranks['Return'] = month_returns
                corr = -stock_ranks.corr(method='spearman').iloc[0][-1]

                month_result['RankCorrelation'] = corr

        strategy_index = strategy_indexes.get(result['StrategyID'])
        if strategy_index is not None:
            results_by_source[result['Context']][strategy_indexes[result['StrategyID']]]['Months'].append(month_result)

    for source, source_results in results_by_source.items():
        for strategy_result in source_results:

            # Sort target months
            strategy_result['Months'] = sorted(strategy_result['Months'], key=lambda x: x['TargetMonth'])

            # Calculate scores
            strategy_result['OverallScore'] = None
            strategy_result['PairScores'] = {}
            strategy_result['Sharpes'] = {}
            strategy_result['MeanReturns'] = {}

            corrs = pd.Series([month_result['RankCorrelation'] for month_result in strategy_result['Months']]).dropna()
            strategy_result['MeanRankCorrelation'] = corrs.mean()
            if len(corrs) > 11:
                strategy_result['MeanRankCorrelationCI'] = calc_bootstrap_confidence_interval(corrs, {'Mean': np.mean})['Mean'].tolist()

            strategy_result['MeanReturnCIs'] = {}
            strategy_result['SharpeCIs'] = {}
            strategy_result['ReturnStdDevs'] = {}

            deltas = [month_result['ReturnDeltas'] for month_result in strategy_result['Months']]
            if None not in deltas:
                deltas = pd.DataFrame(deltas).T

                metrics = factor_calc.calc_long_short_pairs_metrics(deltas)
                strategy_result['OverallScore'] = score_calc.calc_score(metrics)
                for pairs in deltas.index:
                    metrics = factor_calc.calc_long_short_pairs_metrics(deltas.loc[pairs].to_frame())
                    strategy_result['PairScores'][pairs] = score_calc.calc_score(metrics)
                    strategy_result['Sharpes'][pairs] = metrics['Sharpe'].iloc[0]
                    strategy_result['MeanReturns'][pairs] = metrics['Mean'].iloc[0]

                    pair_returns = deltas.loc[pairs].dropna()
                    if len(pair_returns) > 11:
                        ci = calc_bootstrap_confidence_interval(pair_returns, {'Mean': np.mean, 'Sharpe': calc_sharpe})
                        strategy_result['MeanReturnCIs'][pairs] = ci['Mean'].tolist()
                        strategy_result['SharpeCIs'][pairs] = ci['Sharpe'].tolist()

                        strategy_result['ReturnStdDevs'][pairs] = pair_returns.std(ddof=1)

    pairs_range = (
        project_settings['LongShortPairsTarget']-project_settings['LongShortPairsDelta'], 
        project_settings['LongShortPairsTarget']+project_settings['LongShortPairsDelta']
        )

    return (results_by_source, pairs_range)


